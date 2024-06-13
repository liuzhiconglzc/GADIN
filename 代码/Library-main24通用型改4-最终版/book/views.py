import os
import pandas as pd
import json
# TruncWeek
from django.db.models.functions import ExtractMonth,ExtractWeek,TruncMonth
from django.shortcuts import render,get_object_or_404,redirect
from django.urls import  reverse_lazy,reverse
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_protect
from django.views.generic import ListView,DetailView,DeleteView,View,TemplateView
from django.views.generic.edit import CreateView,UpdateView
from django.core.paginator import Paginator
from django.db.models import Q,Sum
from django.http import HttpResponse,HttpResponseRedirect,JsonResponse
from openpyxl import load_workbook
import xlrd
from fillmethod.main_fill import run_model, datafill

from .models import Book,Category,Publisher,UserActivity,Profile,Member,BorrowRecord, UserInfo
from .models import BankInfo, DelConInfo, BankInfoDelMark, ModelParameter
from django.apps import apps
from django.conf import settings
from django.core.paginator import EmptyPage, PageNotAnInteger
from django.core import serializers
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Count
from django.contrib.auth.models import User,Group
from django.contrib.auth.decorators import login_required,user_passes_test,permission_required
from django.utils.decorators import method_decorator

from django.contrib.auth.mixins import LoginRequiredMixin 
from django.contrib.messages.views import SuccessMessageMixin
from django.core.files.storage import FileSystemStorage
from django.contrib.messages.views import messages
from django.views.decorators.csrf import csrf_exempt
from .forms import BookCreateEditForm,PubCreateEditForm,MemberCreateEditForm,ProfileForm,BorrowRecordCreateForm
from .forms import UserCreateEditForm, UploadBankForm, BankCreateEditForm
# from .utils import get_n_days_ago,create_clean_dir,change_col_format
from util.useful import get_n_days_ago,create_clean_dir,change_col_format
from .groups_permissions import check_user_group,user_groups,check_superuser,SuperUserRequiredMixin,allowed_groups
from .custom_filter import get_item
from datetime import date,timedelta,datetime

from django.forms.models import model_to_dict
from django.core.paginator import Paginator
from django.contrib.contenttypes.models import ContentType
from comment.models import Comment
from comment.forms import CommentForm
from notifications.signals import notify
from .notification import send_notification
import logging
import pandas as pd
import numpy as np
from book.utils_df import normalization
from .models import BankInfoFilled, FilledParameter, ParameterInfo
from django.contrib.auth import authenticate
from .models import YhzxInfo, YhzxInfoDelMark, YhzxInfoFilled
from .forms import YhzxCreateEditForm, GeneralCreateEditForm
from itertools import chain
from .models import GeneralInfo, GeneralTableInfo, GeneralMaskInfo, GeneralFilledInfo
import xlwt
import xlsxwriter

logger = logging.getLogger(__name__)

# todo hmf app的视图函数处理文件，url对应uirs.py文件

TODAY=get_n_days_ago(0,"%Y%m%d")
PAGINATOR_NUMBER = 5
allowed_models = ['Category','Publisher','Book','Member','UserActivity','BorrowRecord']



# HomePage

class HomeView(LoginRequiredMixin,TemplateView):
    login_url = 'login'
    template_name = "index.html"
    context={}

  
    # users = User.objects.all()
    # for user in users:
    #     print(user.getUserInfoView_username(),user.is_superuser)

    # def get(self,request, *args, **kwargs):
    #
    #     book_count = Book.objects.aggregate(Sum('quantity'))['quantity__sum']
    #
    #     data_count = {"book":book_count,
    #                 "member":Member.objects.all().count(),
    #                 "category":Category.objects.all().count(),
    #                 "publisher":Publisher.objects.all().count(),}
    #
    #     user_activities= UserActivity.objects.order_by("-created_at")[:5]
    #     user_avatar = { e.created_by:Profile.objects.get(user__username=e.created_by).profile_pic.url for e in user_activities}
    #     short_inventory =Book.objects.order_by('quantity')[:5]
    #
    #     current_week = date.today().isocalendar()[1]
    #     new_members = Member.objects.order_by('-created_at')[:5]
    #     new_members_thisweek = Member.objects.filter(created_at__week=current_week).count()
    #     lent_books_thisweek = BorrowRecord.objects.filter(created_at__week=current_week).count()
    #
    #     books_return_thisweek = BorrowRecord.objects.filter(end_day__week=current_week)
    #     number_books_return_thisweek = books_return_thisweek.count()
    #     new_closed_records = BorrowRecord.objects.filter(open_or_close=1).order_by('-closed_at')[:5]
    #
    #     self.context['data_count']=data_count
    #     self.context['recent_user_activities']=user_activities
    #     self.context['user_avatar']=user_avatar
    #     self.context['short_inventory']=short_inventory
    #     self.context['new_members']=new_members
    #     self.context['new_members_thisweek']=new_members_thisweek
    #     self.context['lent_books_thisweek']=lent_books_thisweek
    #     self.context['books_return_thisweek']=books_return_thisweek
    #     self.context['number_books_return_thisweek']=number_books_return_thisweek
    #     self.context['new_closed_records']=new_closed_records
    #
    #     return render(request, self.template_name, self.context)

# Global Serch
@login_required(login_url='login')
def global_serach(request):
    search_value = request.POST.get('global_search')
    if search_value =='':
        return HttpResponseRedirect("/")

    r_category = Category.objects.filter(Q(name__icontains=search_value))
    r_publisher = Publisher.objects.filter(Q(name__icontains=search_value)|Q(contact__icontains=search_value))
    r_book = Book.objects.filter(Q(author__icontains=search_value)|Q(title__icontains=search_value))
    r_member = Member.objects.filter(Q(name__icontains=search_value)|Q(card_number__icontains=search_value)|Q(phone_number__icontains=search_value))
    r_borrow = BorrowRecord.objects.filter(Q(borrower__icontains=search_value)|Q(borrower_card__icontains=search_value)|Q(book__icontains=search_value))

   
    context={
        'categories':r_category,
        'publishers':r_publisher,
        'books':r_book,
        'members':r_member,
        'records':r_borrow,
    }
    # todo 这是返回到界面
    return render(request, 'book/global_search.html',context=context)
    # return render(request, 'book/userlist.html',context=context)


# Chart
class ChartView(LoginRequiredMixin,TemplateView):
    template_name = "book/charts.html"
    login_url = 'login'
    context={}

    def get(self,request, *args, **kwargs):

        top_5_book= Book.objects.order_by('-quantity')[:5].values_list('title','quantity')
        top_5_book_titles = [b[0] for b in top_5_book ]
        top_5_book__quantities = [b[1] for b in top_5_book ]
        # print(top_5_book_titles,top_5_book__quantities)

        top_borrow = Book.objects.order_by('-total_borrow_times')[:5].values_list('title','total_borrow_times')
        top_borrow_titles = [b[0] for b in top_borrow ]
        top_borrow_times = [b[1] for b in top_borrow ]

        r_open = BorrowRecord.objects.filter(open_or_close=0).count()
        r_close = BorrowRecord.objects.filter(open_or_close=1).count()
        
        m = Member.objects.annotate(month=TruncMonth('created_at')).values('month').annotate(c=Count('id'))
        months_member = [e['month'].strftime("%m/%Y") for e  in m]
        count_monthly_member= [e['c'] for e in m] 

       
        self.context['top_5_book_titles']=top_5_book_titles
        self.context['top_5_book__quantities']=top_5_book__quantities
        self.context['top_borrow_titles']=top_borrow_titles
        self.context['top_borrow_times']=top_borrow_times
        self.context['r_open']=r_open
        self.context['r_close']=r_close
        self.context['months_member']=months_member
        self.context['count_monthly_member']=count_monthly_member
       

        return render(request, self.template_name, self.context)

# Book
class BookListView(LoginRequiredMixin,ListView):
    login_url = 'login'
    model=Book
    context_object_name = 'books'
    template_name = 'book/book_list.html'
    search_value=""
    order_field="-updated_at"

    def get_queryset(self):
        search =self.request.GET.get("search") 
        order_by=self.request.GET.get("orderby")

        if order_by:
            all_books = Book.objects.all().order_by(order_by)
            self.order_field=order_by
        else:
            all_books = Book.objects.all().order_by(self.order_field)

        if search:
            all_books = all_books.filter(
                Q(title__icontains=search)|Q(author__icontains=search)
            )
            self.search_value=search
        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)
        return books

    def get_context_data(self, *args, **kwargs):
        context = super(BookListView, self).get_context_data(*args, **kwargs)
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()
        return context

class BookDetailView(LoginRequiredMixin,DetailView):
    model = Book
    context_object_name = 'book'
    template_name = 'book/book_detail.html'
    login_url = 'login'
    comment_form = CommentForm()
    
    # def get_object(self, queryset=None):
    #     obj = super(BookDetailView, self).get_object(queryset=queryset)
    #     return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_book_name = self.get_object().title
        logger.info(f'Book  <<{current_book_name}>> retrieved from db')
        comments = Comment.objects.filter(book=self.get_object().id)
        related_records = BorrowRecord.objects.filter(book=current_book_name)
        context['related_records'] = related_records
        context['comments'] = comments
        context['comment_form'] = self.comment_form
        return context

class BookCreateView(LoginRequiredMixin,CreateView):
    model=Book
    login_url = 'login'
    form_class=BookCreateEditForm    
    template_name='book/book_create.html'

    def post(self,request, *args, **kwargs):
        super(BookCreateView,self).post(request)
        new_book_name = request.POST['title']
        messages.success(request, f"New Book << {new_book_name} >> Added")
        UserActivity.objects.create(created_by=self.request.user.username,
                                    target_model=self.model.__name__,
                                    detail =f"Create {self.model.__name__} << {new_book_name} >>")
        return redirect('book_list')

class UserCreateView(LoginRequiredMixin,CreateView):
    model=UserInfo
    login_url = 'login'
    form_class=UserCreateEditForm
    template_name='book/user_create.html'

    def post(self,request, *args, **kwargs):
        super(UserCreateView,self).post(request)
        new_user_name = request.POST['username']
        messages.success(request, f"新用户<< {new_user_name} >>已添加  ")
        UserActivity.objects.create(created_by=self.request.user.username,
                                    target_model=self.model.__name__,
                                    detail =f"Create {self.model.__name__} << {new_user_name} >>")
        User.objects.create(username=new_user_name,
                            password="pbkdf2_sha256$100000$Sy6mTZBxqUrk$Wek+897n5GriLAKOdzlQ3qbUhICuHXAwC9822rrcwjQ=")
        return redirect('user_info')

class BookUpdateView(LoginRequiredMixin,UpdateView):
    model = Book
    login_url = 'login'
    form_class=BookCreateEditForm
    template_name = 'book/book_update.html'

    def post(self, request, *args, **kwargs):
        current_book = self.get_object()
        current_book.updated_by=self.request.user.username
        current_book.save(update_fields=['updated_by'])
        UserActivity.objects.create(created_by=self.request.user.username,
            operation_type="warning",
            target_model=self.model.__name__,
            detail =f"Update {self.model.__name__} << {current_book.title} >>")
        return super(BookUpdateView, self).post(request, *args, **kwargs)

    def form_valid(self, form):
      title=form.cleaned_data['title']      
      messages.warning(self.request, f"Update << {title} >> success")
      return super().form_valid(form)

class UserUpdateView(LoginRequiredMixin,UpdateView):
    model = UserInfo
    login_url = 'login'
    form_class=UserCreateEditForm
    template_name = 'book/user_update.html'

    def post(self, request, *args, **kwargs):
        current_book = self.get_object()
        current_book.updated_by=self.request.user.username
        # todo 用来指定更新的字段
        # current_book.save(update_fields=['updated_by'])
        current_book.save(update_fields=['username'])
        UserActivity.objects.create(created_by=self.request.user.username,
            operation_type="warning",
            target_model=self.model.__name__,
            detail =f"Update {self.model.__name__} << {current_book.username} >>")
        return super(UserUpdateView, self).post(request, *args, **kwargs)

    def form_valid(self, form):
      title=form.cleaned_data['username']
      messages.warning(self.request, f"Update << {title} >> success")
      return super().form_valid(form)


# todo 银行信息更新
class BankUpdateView(LoginRequiredMixin,UpdateView):

    #  todo
    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_name = fill_name[0]['fill_name']

    model = GeneralInfo
    form_class = GeneralCreateEditForm
    #
    # if fill_name == "yhzx":
    #     model = YhzxInfo
    #     form_class = YhzxCreateEditForm
    # else:
    #     model = BankInfo
    #     form_class = BankCreateEditForm

    # model = BankInfo
    login_url = 'login'
    # form_class= BankCreateEditForm
    template_name = 'book/bank_update.html'

    def post(self, request, *args, **kwargs):
        current_book = self.get_object()
        current_book.updated_by=self.request.user.username
        # todo 用来指定更新的字段
        # current_book.save(update_fields=['updated_by'])
        current_book.save()
        # todo 类似日志文件的记录
        UserActivity.objects.create(created_by=self.request.user.username,
            operation_type="warning",
            target_model=self.model.__name__,
            detail =f"Update {self.model.__name__} << {current_book} >>")
        return super(BankUpdateView, self).post(request, *args, **kwargs)

    def form_valid(self, form):
      title=form.cleaned_data['col_2']
      messages.warning(self.request, f"Update << {title} >> success")
      return super().form_valid(form)



class BookDeleteView(LoginRequiredMixin,View):
    login_url = 'login'
    def get(self,request,*args,**kwargs):
        book_pk=kwargs["pk"]
        delete_book=Book.objects.get(pk=book_pk)
        model_name = delete_book.__class__.__name__
        messages.error(request, f"Book << {delete_book.title} >> Removed")
        delete_book.delete()
        UserActivity.objects.create(created_by=self.request.user.username,
            operation_type="danger",
            target_model=model_name,
            detail =f"Delete {model_name} << {delete_book.title} >>")
        return HttpResponseRedirect(reverse("book_list"))

# todo 用户删除 hmf
class UserDeleteView(LoginRequiredMixin,View):
    login_url = 'login'
    def get(self,request,*args,**kwargs):
        user_pk=kwargs["pk"]
        delete_user=UserInfo.objects.get(pk=user_pk)
        model_name = delete_user.__class__.__name__
        messages.error(request, f"用户 << {delete_user.username} >> 已删除")
        delete_user.delete()
        delete_also = User.objects.get(pk=user_pk)
        delete_also.delete()
        UserActivity.objects.create(created_by=self.request.user.username,
            operation_type="danger",
            target_model=model_name,
            detail =f"Delete {model_name} << {delete_user.username} >>")
        # todo 同时删除 auth_user 表中数据
        # auth_u = User.objects.get(username=delete_user.username)
        # if auth_u is not None:
        #     auth_u.delete()
        return HttpResponseRedirect(reverse("user_info"))

# todo 银行 hmf
class BankDeleteView(LoginRequiredMixin,View):
    login_url = 'login'
    # fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    # fill_name = fill_name[0]['fill_name']

    def get(self,request,*args,**kwargs):
        user_pk=kwargs["pk"]
        delete_user = GeneralInfo.objects.get(pk=user_pk)
        # if self.fill_name == "yhzx":
        #     delete_user = YhzxInfo.objects.get(pk=user_pk)
        # else:
        #     delete_user = BankInfo.objects.get(pk=user_pk)
        model_name = delete_user.__class__.__name__
        # messages.error(request, f"公司： << {delete_user.company_name} >> 已删除")
        messages.error(request, f"已删除")
        delete_user.delete()
        UserActivity.objects.create(created_by=self.request.user.username,
            operation_type="danger",
            target_model=model_name,
            detail =f"Delete {model_name} <<  >>")
        # todo 同时删除 auth_user 表中数据
        # auth_u = User.objects.get(username=delete_user.username)
        # auth_u.delete()
        return HttpResponseRedirect(reverse("upload_bank"))

# Categorty

class CategoryListView(LoginRequiredMixin,ListView):
    login_url = 'login'
    model=Category
    context_object_name = 'categories'
    template_name = 'book/category_list.html'
    count_total = 0
    search_value = ''
    order_field="-created_at"


    def get_queryset(self):
        search =self.request.GET.get("search")  
        order_by=self.request.GET.get("orderby")
        if order_by:
            all_categories = Category.objects.all().order_by(order_by)
            self.order_field=order_by
        else:
            all_categories = Category.objects.all().order_by(self.order_field)
        if search:
            all_categories = all_categories.filter(
                Q(name__icontains=search)  
            )
            self.search_value=search

        self.count_total = all_categories.count()
        paginator = Paginator(all_categories, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        categories = paginator.get_page(page)
        return categories

    def get_context_data(self, *args, **kwargs):
        context = super(CategoryListView, self).get_context_data(*args, **kwargs)
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()
        return context

class CategoryCreateView(LoginRequiredMixin,CreateView):
    login_url = 'login'
    model=Category
    fields=['name']
    template_name='book/category_create.html'
    success_url = reverse_lazy('category_list')

    def form_valid(self, form):
        new_cat = form.save(commit=False)
        new_cat.save()
        send_notification(self.request.user,new_cat,verb=f'Add New Category << {new_cat.name} >>')
        logger.info(f'{self.request.user} created Category {new_cat.name}')
        UserActivity.objects.create(created_by=self.request.user.username,
                                    target_model=self.model.__name__,
                                    detail =f"Create {self.model.__name__} << {new_cat.name} >>")
        return super(CategoryCreateView, self).form_valid(form)



class CategoryDeleteView(LoginRequiredMixin,View):
    login_url = 'login'

    def get(self,request,*args,**kwargs):
        cat_pk=kwargs["pk"]
        delete_cat=Category.objects.get(pk=cat_pk)
        model_name = delete_cat.__class__.__name__
        messages.error(request, f"Category << {delete_cat.name} >> Removed")
        delete_cat.delete()
        send_notification(self.request.user,delete_cat,verb=f'Delete Category << {delete_cat.name} >>')
        UserActivity.objects.create(created_by=self.request.user.username,
                            operation_type="danger",
                            target_model=model_name,
                            detail =f"Delete {model_name} << {delete_cat.name} >>")

        logger.info(f'{self.request.user} delete Category {delete_cat.name}')

        return HttpResponseRedirect(reverse("category_list"))


# Publisher 

class PublisherListView(LoginRequiredMixin,ListView):
    login_url = 'login'
    model=Publisher
    context_object_name = 'publishers'
    template_name = 'book/publisher_list.html'
    count_total = 0
    search_value = ''
    order_field="-created_at"

    def get_queryset(self):
        search =self.request.GET.get("search")  
        order_by=self.request.GET.get("orderby")
        if order_by:
            all_publishers = Publisher.objects.all().order_by(order_by)
            self.order_field=order_by
        else:
            all_publishers = Publisher.objects.all().order_by(self.order_field)
        if search:
            all_publishers = all_publishers.filter(
                Q(name__icontains=search) | Q(city__icontains=search) | Q(contact__icontains=search)
            )
        else:
            search = ''
        self.search_value=search
        self.count_total = all_publishers.count()
        paginator = Paginator(all_publishers, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        publishers = paginator.get_page(page)
        return publishers

    def get_context_data(self, *args, **kwargs):
        context = super(PublisherListView, self).get_context_data(*args, **kwargs)
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field  
        context['objects'] = self.get_queryset()      
        return context

class PublisherCreateView(LoginRequiredMixin,CreateView):
    model=Publisher
    login_url = 'login'
    form_class=PubCreateEditForm
    template_name='book/publisher_create.html'
    success_url = reverse_lazy('publisher_list')


    def form_valid(self,form):
        new_pub = form.save(commit=False)
        new_pub.save()
        messages.success(self.request, f"New Publisher << {new_pub.name} >> Added")
        send_notification(self.request.user,new_pub,verb=f'Add New Publisher << {new_pub.name} >>')
        logger.info(f'{self.request.user} created Publisher {new_pub.name}')

        UserActivity.objects.create(created_by=self.request.user.username,
                                    target_model=self.model.__name__,
                                    detail =f"Create {self.model.__name__} << {new_pub.name} >>")
        return super(PublisherCreateView, self).form_valid(form)

    # def post(self,request, *args, **kwargs):
    #     super(PublisherCreateView,self).post(request)
    #     new_publisher_name = request.POST['name']
    #     messages.success(request, f"New Publisher << {new_publisher_name} >> Added")
    #     UserActivity.objects.create(created_by=self.request.user.username,
    #                                 target_model=self.model.__name__,
    #                                 detail =f"Create {self.model.__name__} << {new_publisher_name} >>")
    #     return redirect('publisher_list')

class PublisherUpdateView(LoginRequiredMixin,UpdateView):
    model=Publisher
    login_url = 'login'
    form_class=PubCreateEditForm
    template_name = 'book/publisher_update.html'

    def post(self, request, *args, **kwargs):
        current_pub = self.get_object()
        current_pub.updated_by=self.request.user.username
        current_pub.save(update_fields=['updated_by'])
        UserActivity.objects.create(created_by=self.request.user.username,
                                    operation_type="warning",
                                    target_model=self.model.__name__,
                                    detail =f"Update {self.model.__name__} << {current_pub.name} >>")
        return super(PublisherUpdateView, self).post(request, *args, **kwargs)

    def form_valid(self, form):
        title=form.cleaned_data['name']      
        messages.warning(self.request, f"Update << {title} >> success")
        return super().form_valid(form)

class PublisherDeleteView(LoginRequiredMixin,View):
    login_url = 'login'

    def get(self,request,*args,**kwargs):
        pub_pk=kwargs["pk"]
        delete_pub=Publisher.objects.get(pk=pub_pk)
        model_name = delete_pub.__class__.__name__
        messages.error(request, f"Publisher << {delete_pub.name} >> Removed")
        delete_pub.delete()
        send_notification(self.request.user,delete_pub,verb=f'Delete Publisher << {delete_pub.name} >>')
        logger.info(f'{self.request.user} delete Publisher {delete_pub.name}')
        UserActivity.objects.create(created_by=self.request.user.username,
                    operation_type="danger",
                    target_model=model_name,
                    detail =f"Delete {model_name} << {delete_pub.name} >>")
        return HttpResponseRedirect(reverse("publisher_list"))


# User Logs  
# @method_decorator(user_passes_test(lambda u: u.has_perm("book.view_useractivity")), name='dispatch')
@method_decorator(allowed_groups(group_name=['logs']), name='dispatch')
class ActivityListView(LoginRequiredMixin,ListView):

    login_url = 'login'
    model= UserActivity
    context_object_name = 'activities'
    template_name = 'book/user_activity_list.html'
    count_total = 0
    search_value=''
    created_by=''
    order_field="-created_at"
    all_users = User.objects.values()
    user_list = [x['username'] for x in all_users] 

    # def dispatch(self, *args, **kwargs):
    #     return super(ActivityListView, self).dispatch(*args, **kwargs)

    def get_queryset(self):
        data = self.request.GET.copy()
        search =self.request.GET.get("search")
        filter_user=self.request.GET.get("created_by") 

        all_activities = UserActivity.objects.all()

        if filter_user:
            self.created_by = filter_user
            all_activities = all_activities.filter(created_by=self.created_by)

        if search:
            self.search_value = search
            all_activities = all_activities.filter(Q(target_model__icontains=search))

        self.search_value=search
        self.count_total = all_activities.count()
        paginator = Paginator(all_activities,PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        try:
            response = paginator.get_page(page)
        except PageNotAnInteger:
            response = paginator.get_page(1)
        except EmptyPage:
            response = paginator.get_page(paginator.num_pages)
        return response

    
    def get_context_data(self, *args, **kwargs):
        context = super(ActivityListView, self).get_context_data(*args, **kwargs)
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['user_list']= self.user_list
        context['created_by'] = self.created_by
        return context


# @method_decorator(user_passes_test(lambda u: u.has_perm("book.delete_useractivity")), name='dispatch')
@method_decorator(allowed_groups(group_name=['logs']), name='dispatch')
class ActivityDeleteView(LoginRequiredMixin,View):

    login_url = 'login'

    def get(self,request,*args,**kwargs):
        
        log_pk=kwargs["pk"]
        delete_log=UserActivity.objects.get(pk=log_pk)
        messages.error(request, f"Activity Removed")
        delete_log.delete()

        return HttpResponseRedirect(reverse("user_activity_list"))


# Membership
class MemberListView(LoginRequiredMixin,ListView):
    login_url = 'login'
    model= Member
    context_object_name = 'members'
    template_name = 'book/member_list.html'
    count_total = 0
    search_value = ''
    order_field="-updated_at"

    def get_queryset(self):
        search =self.request.GET.get("search")  
        order_by=self.request.GET.get("orderby")
        if order_by:
            all_members = Member.objects.all().order_by(order_by)
            self.order_field=order_by
        else:
            all_members = Member.objects.all().order_by(self.order_field)
        if search:
            all_members = all_members.filter(
                Q(name__icontains=search) |  Q(card_number__icontains=search)
            )
        else:
            search = ''
        self.search_value=search
        self.count_total = all_members.count()
        paginator = Paginator(all_members, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        members = paginator.get_page(page)
        return members

    def get_context_data(self, *args, **kwargs):
        context = super(MemberListView, self).get_context_data(*args, **kwargs)
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()
        return context

class MemberCreateView(LoginRequiredMixin,CreateView):
    model=Member
    login_url = 'login'
    form_class=MemberCreateEditForm
    template_name='book/member_create.html'

    def post(self,request, *args, **kwargs):
        super(MemberCreateView,self).post(request)
        new_member_name = request.POST['name']
        messages.success(request, f"New Member << {new_member_name} >> Added")
        UserActivity.objects.create(created_by=self.request.user.username,
                                    target_model=self.model.__name__,
                                    detail =f"Create {self.model.__name__} << {new_member_name} >>")
        return redirect('member_list')

    def form_valid(self, form):
        self.object = form.save()
        self.object.created_by = self.request.user.username
        self.object.save(update_fields=['created_by'])
        send_notification(self.request.user,self.object,f'Add new memeber {self.object.name}')
    
        return HttpResponseRedirect(self.get_success_url())


    # def form_valid(self, form):
    #     response = super(CourseCreate, self).form_valid(form)
    #     # do something with self.object
    #     return response

class MemberUpdateView(LoginRequiredMixin,UpdateView):
    model = Member
    login_url = 'login'
    form_class=MemberCreateEditForm
    template_name = 'book/member_update.html'

    def post(self, request, *args, **kwargs):
        current_member = self.get_object()
        current_member.updated_by=self.request.user.username
        current_member.save(update_fields=['updated_by'])
        UserActivity.objects.create(created_by=self.request.user.username,
            operation_type="warning",
            target_model=self.model.__name__,
            detail =f"Update {self.model.__name__} << {current_member.name} >>")
        return super(MemberUpdateView, self).post(request, *args, **kwargs)

    def form_valid(self, form):
        member_name=form.cleaned_data['name']      
        messages.warning(self.request, f"Update << {member_name} >> success")
        return super().form_valid(form)

class MemberDeleteView(LoginRequiredMixin,View):
    login_url = 'login'

    def get(self,request,*args,**kwargs):
        member_pk=kwargs["pk"]
        delete_member=Member.objects.get(pk=member_pk)
        model_name = delete_member.__class__.__name__
        messages.error(request, f"Member << {delete_member.name} >> Removed")
        delete_member.delete()
        send_notification(self.request.user,delete_member,f'Delete member {delete_member.name} ')


        UserActivity.objects.create(created_by=self.request.user.username,
                    operation_type="danger",
                    target_model=model_name,
                    detail =f"Delete {model_name} << {delete_member.name} >>")
        return HttpResponseRedirect(reverse("member_list"))

class MemberDetailView(LoginRequiredMixin,DetailView):
    model = Member
    context_object_name = 'member'
    template_name = 'book/member_detail.html'
    login_url = 'login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_member_name = self.get_object().name
        related_records = BorrowRecord.objects.filter(borrower=current_member_name)
        context['related_records'] = related_records
        context["card_number"] = str(self.get_object().card_id)[:8]
        return context


# Profile View

class ProfileDetailView(LoginRequiredMixin,DetailView):
    model = Profile
    context_object_name = 'profile'
    template_name = 'profile/profile_detail.html'
    login_url = 'login'


    def get_context_data(self, *args, **kwargs):
        current_user= get_object_or_404(Profile,pk=self.kwargs['pk'])
        # current_user= Profile.get(pk=kwargs['pk'])
        context = super(ProfileDetailView, self).get_context_data(*args, **kwargs)
        context['current_user'] = current_user
        return context

class ProfileCreateView(LoginRequiredMixin,CreateView):
    model = Profile
    template_name = 'profile/profile_create.html'
    login_url = 'login'
    form_class= ProfileForm

    def form_valid(self,form) -> HttpResponse:
        form.instance.user = self.request.user
        return super().form_valid(form)

class ProfileUpdateView(LoginRequiredMixin,UpdateView):
    model = Profile
    login_url = 'login'
    form_class=ProfileForm
    template_name = 'profile/profile_update.html'

# Borrow Records 

class BorrowRecordCreateView(LoginRequiredMixin,CreateView):
    model = BorrowRecord
    template_name = 'borrow_records/create.html'
    form_class=BorrowRecordCreateForm
    login_url = 'login'

    

    def get_form(self):
        form = super().get_form()
        return form

    def form_valid(self, form):
        selected_member= get_object_or_404(Member,name = form.cleaned_data['borrower'] )
        selected_book = Book.objects.get(title=form.cleaned_data['book'])

        # if form.is_valid():
        #     form.save(commit=True)
        #     return HttpResponse("Successfully added the date to database");
        # else:
        #     # The supplied form contained errors - just print them to the terminal.
        #     print(form.errors)

        form.instance.borrower_card = selected_member.card_number
        form.instance.borrower_email = selected_member.email
        form.instance.borrower_phone_number = selected_member.phone_number
        form.instance.created_by = self.request.user.username
        form.instance.start_day = form.cleaned_data['start_day']
        form.instance.end_day = form.cleaned_data['end_day']
        form.save()


        # Change field on Model Book
        selected_book.status=0
        selected_book.total_borrow_times+=1
        selected_book.quantity-=int(form.cleaned_data['quantity'])
        selected_book.save()

        # Create Log 
        borrower_name = selected_member.name
        book_name = selected_book.title

        messages.success(self.request, f" '{borrower_name}' borrowed <<{book_name}>>")
        UserActivity.objects.create(created_by=self.request.user.username,
                                    target_model=self.model.__name__,
                                    detail =f" '{borrower_name}' borrowed <<{book_name}>>")


        return super(BorrowRecordCreateView,self).form_valid(form)

 
    # def post(self,request, *args, **kwargs):

    #     return redirect('record_list')



@login_required(login_url='login')
def auto_member(request):
    if request.is_ajax():
        query = request.GET.get("term", "")
        member_names = Member.objects.filter(name__icontains=query)
        results = []
        for m in member_names:
            results.append(m.name)
        data = json.dumps(results)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@login_required(login_url='login')
def auto_book(request):
    if request.is_ajax():
        query = request.GET.get("term", "")
        book_names = Book.objects.filter(title__icontains=query)
        results = [b.title for b in book_names]
        data = json.dumps(results)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

class BorrowRecordDetailView(LoginRequiredMixin,DetailView):
    model = BorrowRecord
    context_object_name = 'record'
    template_name = 'borrow_records/detail.html'
    login_url = 'login'   

    # def get_queryset(self):
    #     return BorrowRecord.objects.filter(pk=self.kwargs['pk'])

    # Not recommanded
    def get_context_data(self, **kwargs):
        context = super(BorrowRecordDetailView, self).get_context_data(**kwargs)
        related_member = Member.objects.get(name=self.get_object().borrower)
        context['related_member'] = related_member
        return context

class BorrowRecordListView(LoginRequiredMixin,ListView):
    model = BorrowRecord
    template_name = 'borrow_records/list.html'
    login_url = 'login'
    context_object_name = 'records'
    count_total = 0
    search_value = ''
    order_field="-closed_at"

    def get_queryset(self):
        search =self.request.GET.get("search")  
        order_by=self.request.GET.get("orderby")
        if order_by:
            all_records = BorrowRecord.objects.all().order_by(order_by)
            self.order_field=order_by
        else:
            all_records = BorrowRecord.objects.all().order_by(self.order_field)
        if search:
            all_records = BorrowRecord.objects.filter(
                Q(borrower__icontains=search) | Q(book__icontains=search) | Q(borrower_card__icontains=search)
            )
        else:
            search = ''
        self.search_value=search
        self.count_total = all_records.count()
        paginator = Paginator(all_records, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        records = paginator.get_page(page)
        return records

    def get_context_data(self, *args, **kwargs):
        context = super(BorrowRecordListView, self).get_context_data(*args, **kwargs)
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()
        return context

#     todo 用戶信息
class UserInfoView(LoginRequiredMixin,ListView):
    login_url = 'login'
    model = UserInfo
    context_object_name = 'books'
    template_name = 'book/userlist.html'
    search_value = ""
    order_field = "id"

    def get_queryset(self):
        search = self.request.GET.get("search")
        order_by = self.request.GET.get("orderby")
        # a = User.objects.all()

        if order_by:
            all_books = UserInfo.objects.all().order_by(order_by)
            self.order_field = order_by
        else:
            all_books = UserInfo.objects.all().order_by(self.order_field)
        # todo 这是搜索框条件id,username
        if search:
            all_books = all_books.filter(
                Q(id__icontains=search) | Q(username__icontains=search)
            )
            self.search_value = search
        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)
        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(UserInfoView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()

        # username = self.request.user.username
        # # user = authenticate(username=username)
        # # todo 管理员
        # user_info = UserInfo.objects.filter(username=username)
        # if user_info is not None:
        #     return redirect('show')

        return context


class BorrowRecordDeleteView(LoginRequiredMixin,View):
    login_url = 'login'

    def get(self,request,*args,**kwargs):
        record_pk=kwargs["pk"]
        delete_record=BorrowRecord.objects.get(pk=record_pk)
        model_name = delete_record.__class__.__name__
        messages.error(request, f"Record {delete_record.borrower} => {delete_record.book} Removed")
        delete_record.delete()
        UserActivity.objects.create(created_by=self.request.user.username,
                    operation_type="danger",
                    target_model=model_name,
                    detail =f"Delete {model_name} {delete_record.borrower}")
        return HttpResponseRedirect(reverse("record_list"))

class BorrowRecordClose(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):

        close_record = BorrowRecord.objects.get(pk=self.kwargs['pk'])
        close_record.closed_by = self.request.user.username
        close_record.final_status = close_record.return_status
        close_record.delay_days = close_record.get_delay_number_days
        close_record.open_or_close = 1
        close_record.save()
        print(close_record.open_or_close,close_record.final_status,close_record.pk)
        

        borrowed_book = Book.objects.get(title=close_record.book)
        borrowed_book.quantity+=1
        count_record_same_book = BorrowRecord.objects.filter(book=close_record.book).count()
        if count_record_same_book==1:
            borrowed_book.status = 1

        borrowed_book.save()

        model_name = close_record.__class__.__name__
        UserActivity.objects.create(created_by=self.request.user.username,
                    operation_type="info",
                    target_model=model_name,
                    detail =f"Close {model_name} '{close_record.borrower}'=>{close_record.book}")
        return HttpResponseRedirect(reverse("record_list"))


# Data center
@method_decorator(allowed_groups(group_name=['download_data']), name='dispatch')
class DataCenterView(LoginRequiredMixin,TemplateView):
    template_name = 'book/download_data.html'
    login_url = 'login'

    def get(self,request,*args, **kwargs):
        # check_user_group(request.user,"download_data")
        data = {m.objects.model._meta.db_table:
        {"source":pd.DataFrame(list(m.objects.all().values())) ,
          "path":f"{str(settings.BASE_DIR)}/datacenter/{m.__name__}_{TODAY}.csv",
           "file_name":f"{m.__name__}_{TODAY}.csv"} for m in apps.get_models() if m.__name__ in allowed_models}
        
        count_total = {k: v['source'].shape[0] for k,v in data.items()}
        return render(request,self.template_name,context={'model_list':count_total})

@login_required(login_url='login')
@allowed_groups(group_name=['download_data'])
def download_data(request,model_name):
    check_user_group(request.user,"download_data")
            
    download = {m.objects.model._meta.db_table:
        {"source":pd.DataFrame(list(m.objects.all().values())) ,
          "path":f"{str(settings.BASE_DIR)}/datacenter/{m.__name__}_{TODAY}.csv",
           "file_name":f"{m.__name__}_{TODAY}.csv"} for m in apps.get_models() if m.__name__ in allowed_models}

    download[model_name]['source'].to_csv(download[model_name]['path'],index=False,encoding='utf-8')
    download_file=pd.read_csv(download[model_name]['path'],encoding='utf-8')
    response = HttpResponse(download_file,content_type="text/csv")
    response = HttpResponse(open(download[model_name]['path'],'r',encoding='utf-8'),content_type="text/csv")
    response['Content-Disposition'] = f"attachment;filename={download[model_name]['file_name']}"
    return response


    
# Handle Errors

def page_not_found(request, exception):
    context = {}
    response = render(request, "errors/404.html", context=context)
    response.status_code = 404
    return response
    
def server_error(request, exception=None):
    context = {}
    response = render(request, "errors/500.html", context=context)
    response.status_code = 500
    return response
    
def permission_denied(request, exception=None):
    context = {}
    response = render(request, "errors/403.html", context=context)
    response.status_code = 403
    return response
    
def bad_request(request, exception=None):
    context = {}
    response = render(request, "errors/400.html", context=context)
    response.status_code = 400
    return response

# Employees
# @method_decorator(user_passes_test(lambda u: check_superuser(u)), name='dispatch')
class EmployeeView(SuperUserRequiredMixin,ListView):
    login_url = 'login'
    model=User
    context_object_name = 'employees'
    template_name = 'book/employees.html'

    # def get(self, request):
    #     # check_superuser(request.user)
    #     return super(EmployeeView, self).get(self,request)

# @method_decorator(user_passes_test(lambda u: check_superuser(u)), name='dispatch')
class EmployeeDetailView(SuperUserRequiredMixin,DetailView):
    model = User
    context_object_name = 'employee'
    template_name = 'book/employee_detail.html'
    login_url = 'login'


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['groups'] = user_groups
        return context


@user_passes_test(lambda u: u.is_superuser)
@login_required(login_url='login')
def EmployeeUpdate(request,pk):
    # check_superuser(request.user)
    current_user = User.objects.get(pk=pk)
    if request.method == 'POST':
        chosen_groups = [ g for g in user_groups if "on" in request.POST.getlist(g)]
        current_user.groups.clear()
        for each in chosen_groups:
            group = Group.objects.get(name=each)
            current_user.groups.add(group)
        messages.success(request, f"Group for  << {current_user.username} >> has been updated")
        return redirect('employees_detail', pk=pk)



# Notice

class NoticeListView(SuperUserRequiredMixin, ListView):
    context_object_name = 'notices'
    template_name = 'notice_list.html'
    login_url = 'login'

    # 未读通知的查询集
    def get_queryset(self):
        return self.request.user.notifications.unread()


class NoticeUpdateView(SuperUserRequiredMixin,View):
    """Update Status of Notification"""
    # 处理 get 请求
    def get(self, request):
        # 获取未读消息
        notice_id = request.GET.get('notice_id')
        # 更新单条通知
        if notice_id:
            request.user.notifications.get(id=notice_id).mark_as_read()
            return redirect('category_list')
        # 更新全部通知
        else:
            request.user.notifications.mark_all_as_read()
            return redirect('notice_list')



# todo 加载excel文件并保存到数据库------不能用
class UploadView(LoginRequiredMixin, CreateView):
    model = Profile
    template_name = 'book/load_excel.html'
    login_url = 'login'
    form_class = UserCreateEditForm

    def form_valid(self, form) -> HttpResponse:
        form.instance.user = self.request.user
        return super().form_valid(form)

    def upload_bank(request):
        if request.method == "POST":  # 验证POST
            uf = UploadBankForm(request.POST, request.FILES)  # .post是获取post返回字段，.FILES是获取返回的文件
            print(uf)
            print(request.FILES['uploadfile'])
            print('-----------')
            if uf.is_valid():  # 判断前台返回的表单是否为有效的类型
                wb = load_workbook(filename=request.FILES['uploadfile'])
                print(wb)
                ws = wb.get_sheet_names()
                ws = wb.get_sheet_by_name(ws[0])
                max_row = ws.max_row
                for row in range(2, max_row + 1):
                    # 获取表单元素
                    izcfz_id = ws.cell(row=row, column=1).value
                    company_name = ws.cell(row=row, column=2).value
                    type = ws.cell(row=row, column=3).value
                    bbd_url = ws.cell(row=row, column=4).value
                    bbd_type = ws.cell(row=row, column=5).value
                    cash_central_bank_funds = ws.cell(row=row, column=6).value
                    sellable_assetset = ws.cell(row=row, column=7).value
                    risk_preparation = ws.cell(row=row, column=8).value
                    interest_payable = ws.cell(row=row, column=9).value
                    paid_in_capital = ws.cell(row=row, column=10).value
                    fixed_assets = ws.cell(row=row, column=11).value
                    total_assets = ws.cell(row=row, column=12).value
                    capital_reserves = ws.cell(row=row, column=13).value
                    # 写入数据库
                    upload_bank = BankInfo()
                    upload_bank.izcfz_id = izcfz_id
                    upload_bank.company_name = company_name
                    upload_bank.type = type
                    upload_bank.bbd_url = bbd_url
                    upload_bank.bbd_type = bbd_type
                    upload_bank.cash_central_bank_funds = cash_central_bank_funds
                    upload_bank.sellable_assetset = sellable_assetset
                    upload_bank.risk_preparation = risk_preparation
                    upload_bank.interest_payable = interest_payable
                    upload_bank.paid_in_capital = paid_in_capital
                    upload_bank.fixed_assets = fixed_assets
                    upload_bank.total_assets = total_assets
                    upload_bank.capital_reserves = capital_reserves

                    upload_bank.save()
                return HttpResponse('upload ok!')
        else:
            uf = UploadBankForm()
        return render(request, 'book/load_excel.html', {'uf': uf})


# todo 方式1：直接跳转界面 不用
def uploadexcel(request):
    return render(request, "book/load_excel.html")

# todo 方式2：尝试新的界面
class BankInfoView(LoginRequiredMixin,ListView):

    login_url = 'login'
    model = GeneralInfo
    template_name = 'book/load_excel_general.html'

    context_object_name = 'books'
    search_value = ""
    order_field = "id"

    condition1 = ""
    condition2 = ""
    condition3 = ""
    condition4 = ""
    condition5 = ""
    condition6 = ""
    condition7 = ""
    condition8 = ""
    condition9 = ""
    condition10 = ""
    condition11 = ""
    condition12 = ""
    condition13 = ""
    condition14 = ""
    condition15 = ""
    condition16 = ""
    condition17 = ""
    condition18 = ""
    condition19 = ""
    condition20 = ""

    def get_queryset(self):
        # todo 加一个判断，：当下正在操作的表
        fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
        fill_nameed = fill_name[0]['fill_name']

        self.condition1 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        self.condition2 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        self.condition3 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        self.condition4 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        self.condition5 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        self.condition6 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        self.condition7 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        self.condition8 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        self.condition9 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        self.condition10 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        self.condition11 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        self.condition12 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        self.condition13 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        self.condition14 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        self.condition15 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        self.condition16 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        self.condition17 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        self.condition18 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        self.condition19 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        self.condition20 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        search = self.request.GET.get("search")
        status_con1 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        status_con2 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        status_con3 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        status_con4 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        status_con5 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        status_con6 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        status_con7 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        status_con8 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        status_con9 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        status_con10 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        status_con11 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        status_con12 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        status_con13 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        status_con14 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        status_con15 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        status_con16 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        status_con17 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        status_con18 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        status_con19 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        status_con20 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        if status_con1 == "2":
            self.condition1 = ""
        if status_con2 == "2":
            self.condition2 = ""
        if status_con3 == "2":
            self.condition3 = ""
        if status_con4 == "2":
            self.condition4 = ""
        if status_con5 == "2":
            self.condition5 = ""
        if status_con6 == "2":
            self.condition6 = ""
        if status_con7 == "2":
            self.condition7 = ""
        if status_con8 == "2":
            self.condition8 = ""
        if status_con9 == "2":
            self.condition9 = ""
        if status_con10 == "2":
            self.condition10 = ""
        if status_con11 == "2":
            self.condition11 = ""
        if status_con12 == "2":
            self.condition12 = ""
        if status_con13 == "2":
            self.condition13 = ""
        if status_con14 == "2":
            self.condition14 = ""
        if status_con15 == "2":
            self.condition15 = ""
        if status_con16 == "2":
            self.condition16 = ""
        if status_con17 == "2":
            self.condition17 = ""
        if status_con18 == "2":
            self.condition18 = ""
        if status_con19 == "2":
            self.condition19 = ""
        if status_con20 == "2":
            self.condition20 = ""
        all_books = GeneralInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)

        if search:
            # all_books = all_books.filter(
            #     Q(id__icontains=search) | Q(username__icontains=search)
            # )
            # todo 暂不做处理
            if search in "中小板":
                all_books.filter(col_3__contains="中小板").update(col_3=1)
            elif search in "上证A股":
                all_books.filter(col_3__contains="上证A股").update(col_3=2)
            elif search in "新三板":
                all_books.filter(col_3__contains="新三板").update(col_3=3)
            else:
                None

            all_books = GeneralInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)
            self.search_value = search

        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)

        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(BankInfoView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()

        context['condition1'] = self.condition1
        context['condition2'] = self.condition2
        context['condition3'] = self.condition3
        context['condition4'] = self.condition4
        context['condition5'] = self.condition5
        context['condition6'] = self.condition6
        context['condition7'] = self.condition7
        context['condition8'] = self.condition8
        context['condition9'] = self.condition9
        context['condition10'] = self.condition10
        context['condition11'] = self.condition11
        context['condition12'] = self.condition12
        context['condition13'] = self.condition13
        context['condition14'] = self.condition14
        context['condition15'] = self.condition15
        context['condition16'] = self.condition16
        context['condition17'] = self.condition17
        context['condition18'] = self.condition18
        context['condition19'] = self.condition19
        context['condition20'] = self.condition20
        return context


# todo 冗余属性删除功能
class BankInfoDelView(LoginRequiredMixin,ListView):

    model = GeneralInfo
    template_name = 'book/load_excel_del_general.html'
    login_url = 'login'
    context_object_name = 'books'
    search_value = ""
    order_field = "id"

    message_do = 1
    condition1 = ""
    condition2 = ""
    condition3 = ""
    condition4 = ""
    condition5 = ""
    condition6 = ""
    condition7 = ""
    condition8 = ""
    condition9 = ""
    condition10 = ""
    condition11 = ""
    condition12 = ""
    condition13 = ""
    condition14 = ""
    condition15 = ""
    condition16 = ""
    condition17 = ""
    condition18 = ""
    condition19 = ""
    condition20 = ""

    def get_queryset(self):
        self.message_do = 1
        # todo 加一个判断，：当下正在操作的表
        fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
        fill_nameed = fill_name[0]['fill_name']

        self.condition1 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        self.condition2 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        self.condition3 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        self.condition4 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        self.condition5 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        self.condition6 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        self.condition7 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        self.condition8 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        self.condition9 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        self.condition10 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        self.condition11 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        self.condition12 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        self.condition13 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        self.condition14 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        self.condition15 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        self.condition16 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        self.condition17 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        self.condition18 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        self.condition19 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        self.condition20 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        search = self.request.GET.get("search")

        status_con1 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        status_con2 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        status_con3 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        status_con4 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        status_con5 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        status_con6 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        status_con7 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        status_con8 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        status_con9 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        status_con10 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        status_con11 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        status_con12 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        status_con13 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        status_con14 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        status_con15 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        status_con16 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        status_con17 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        status_con18 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        status_con19 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        status_con20 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        if status_con1 == "2":
            self.condition1 = ""
        if status_con2 == "2":
            self.condition2 = ""
        if status_con3 == "2":
            self.condition3 = ""
        if status_con4 == "2":
            self.condition4 = ""
        if status_con5 == "2":
            self.condition5 = ""
        if status_con6 == "2":
            self.condition6 = ""
        if status_con7 == "2":
            self.condition7 = ""
        if status_con8 == "2":
            self.condition8 = ""
        if status_con9 == "2":
            self.condition9 = ""
        if status_con10 == "2":
            self.condition10 = ""
        if status_con11 == "2":
            self.condition11 = ""
        if status_con12 == "2":
            self.condition12 = ""
        if status_con13 == "2":
            self.condition13 = ""
        if status_con14 == "2":
            self.condition14 = ""
        if status_con15 == "2":
            self.condition15 = ""
        if status_con16 == "2":
            self.condition16 = ""
        if status_con17 == "2":
            self.condition17 = ""
        if status_con18 == "2":
            self.condition18 = ""
        if status_con19 == "2":
            self.condition19 = ""
        if status_con20 == "2":
            self.condition20 = ""

        all_books = GeneralInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)

        # todo 这是搜索框条件，如果有内容执行下列程序
        if search:
            if search == self.condition1:
                self.condition1 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_1="2")
            elif search == self.condition2:
                self.condition2 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_2="2")
            elif search == self.condition3:
                self.condition3 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_3="2")
            elif search == self.condition4:
                self.condition4 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_4="2")
            elif search == self.condition5:
                self.condition5 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_5="2")
            elif search == self.condition6:
                self.condition6 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_6="2")
            elif search == self.condition7:
                self.condition7 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_7="2")
            elif search == self.condition8:
                self.condition8 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_8="2")
            elif search == self.condition9:
                self.condition9 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_9="2")
            elif search == self.condition10:
                self.condition10 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_10="2")
            elif search == self.condition11:
                self.condition11 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_11="2")
            elif search == self.condition12:
                self.condition12 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_12="2")
            elif search == self.condition13:
                self.condition13 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_13="2")
            elif search == self.condition14:
                self.condition14 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_14="2")
            elif search == self.condition15:
                self.condition15 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_15="2")
            elif search == self.condition16:
                self.condition16 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_16="2")
            elif search == self.condition17:
                self.condition17 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_17="2")
            elif search == self.condition18:
                self.condition18 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_18="2")
            elif search == self.condition19:
                self.condition19 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_19="2")
            elif search == self.condition20:
                self.condition20 = ""
                GeneralMaskInfo.objects.filter(table_name=fill_nameed).update(col_20="2")
            else:
                self.message_do = 2

            all_books = GeneralInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)
            self.search_value = search

        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)
        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(BankInfoDelView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()

        # todo 放在上一个方法里会执行 两次
        if self.message_do == 2:
            None
            # messages.warning(self.request, f"没有输入有效属性名称")
        # todo
        context['condition1'] = self.condition1
        context['condition2'] = self.condition2
        context['condition3'] = self.condition3
        context['condition4'] = self.condition4
        context['condition5'] = self.condition5
        context['condition6'] = self.condition6
        context['condition7'] = self.condition7
        context['condition8'] = self.condition8
        context['condition9'] = self.condition9
        context['condition10'] = self.condition10
        context['condition11'] = self.condition11
        context['condition12'] = self.condition12
        context['condition13'] = self.condition13
        context['condition14'] = self.condition14
        context['condition15'] = self.condition15
        context['condition16'] = self.condition16
        context['condition17'] = self.condition17
        context['condition18'] = self.condition18
        context['condition19'] = self.condition19
        context['condition20'] = self.condition20

        return context

# todo 数据归一化功能
class BankInfoNovView(LoginRequiredMixin,ListView):
    login_url = 'login'
    model = BankInfo
    context_object_name = 'books'
    template_name = 'book/load_excel_nov.html'
    search_value = ""
    order_field = "id"
    # todo 新加控制属性
    # DelConInfo.objects.filter(del_status__icontains="完成").delete()
    condition1 = True
    condition2 = True
    err_is = 1
    def get_queryset(self):
        search = self.request.GET.get("search")
        order_by = self.request.GET.get("orderby")

        status_con1 = DelConInfo.objects.values("del_name").first()
        status_con2 = DelConInfo.objects.values("del_status").first()

        if status_con1['del_name'] == "是":
            self.condition1 = False
        if status_con2['del_status'] == "是":
            self.condition2 = False

        if order_by:
            all_books = BankInfo.objects.all().order_by(order_by)
            self.order_field = order_by
        else:
            all_books = BankInfo.objects.all().order_by(self.order_field)
        # todo 这是搜索框条件，如果有内容执行下列程序
        if search:
            # all_books = all_books.filter(
            #     Q(id__icontains=search) | Q(username__icontains=search)
            # )
            # status_del = DelConInfo.objects.values("del_name").last()
            # print(status_del['del_name'])
            if search == "数据来源" and status_con1['del_name'] == "否":
                self.condition1 = False
                DelConInfo.objects.filter(del_name="否").update(del_name="是")
            elif search == "数据类型" and status_con2['del_status'] == "否":
                self.condition2 = False
                DelConInfo.objects.filter(del_status="否").update(del_status="是")
            else:
                self.err_is = 2

            all_books = BankInfo.objects.all().order_by(self.order_field)
            self.search_value = search
        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)
        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(BankInfoNovView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()

        # todo 放在上一个方法里会执行 两次
        # todo
        context['condition1'] = self.condition1
        context['condition2'] = self.condition2

        # # todo
        # if self.err_is == 2:
        #     messages.warning(self.request, "没有输入正确属性")

        # todo
        # if self.condition1 == False and self.condition2 == False:
        #     DelConInfo.objects.create(del_name=self.search_value, del_status="完成")
        return context

# todo 计算缺失率功能
class BankInfoConView(LoginRequiredMixin,ListView):

    login_url = 'login'

    model = GeneralInfo
    context_object_name = 'books'
    template_name = 'book/load_excel_con_general.html'
    search_value = ""
    order_field = "id"
    # todo 新加控制属性
    condition1 = ""
    condition2 = ""
    condition3 = ""
    condition4 = ""
    condition5 = ""
    condition6 = ""
    condition7 = ""
    condition8 = ""
    condition9 = ""
    condition10 = ""
    condition11 = ""
    condition12 = ""
    condition13 = ""
    condition14 = ""
    condition15 = ""
    condition16 = ""
    condition17 = ""
    condition18 = ""
    condition19 = ""
    condition20 = ""

    def get_queryset(self):
        # todo 加一个判断，：当下正在操作的表
        fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
        fill_nameed = fill_name[0]['fill_name']

        self.condition1 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        self.condition2 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        self.condition3 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        self.condition4 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        self.condition5 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        self.condition6 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        self.condition7 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        self.condition8 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        self.condition9 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        self.condition10 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        self.condition11 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        self.condition12 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        self.condition13 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        self.condition14 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        self.condition15 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        self.condition16 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        self.condition17 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        self.condition18 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        self.condition19 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        self.condition20 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        # search = self.request.GET.get("search")

        status_con1 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        status_con2 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        status_con3 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        status_con4 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        status_con5 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        status_con6 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        status_con7 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        status_con8 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        status_con9 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        status_con10 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        status_con11 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        status_con12 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        status_con13 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        status_con14 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        status_con15 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        status_con16 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        status_con17 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        status_con18 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        status_con19 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        status_con20 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        if status_con1 == "2":
            self.condition1 = ""
        if status_con2 == "2":
            self.condition2 = ""
        if status_con3 == "2":
            self.condition3 = ""
        if status_con4 == "2":
            self.condition4 = ""
        if status_con5 == "2":
            self.condition5 = ""
        if status_con6 == "2":
            self.condition6 = ""
        if status_con7 == "2":
            self.condition7 = ""
        if status_con8 == "2":
            self.condition8 = ""
        if status_con9 == "2":
            self.condition9 = ""
        if status_con10 == "2":
            self.condition10 = ""
        if status_con11 == "2":
            self.condition11 = ""
        if status_con12 == "2":
            self.condition12 = ""
        if status_con13 == "2":
            self.condition13 = ""
        if status_con14 == "2":
            self.condition14 = ""
        if status_con15 == "2":
            self.condition15 = ""
        if status_con16 == "2":
            self.condition16 = ""
        if status_con17 == "2":
            self.condition17 = ""
        if status_con18 == "2":
            self.condition18 = ""
        if status_con19 == "2":
            self.condition19 = ""
        if status_con20 == "2":
            self.condition20 = ""
        all_books = GeneralInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)

        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)

        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(BankInfoConView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()

        # todo 放在上一个方法里会执行 两次
        # todo
        context['condition1'] = self.condition1
        context['condition2'] = self.condition2
        context['condition3'] = self.condition3
        context['condition4'] = self.condition4
        context['condition5'] = self.condition5
        context['condition6'] = self.condition6
        context['condition7'] = self.condition7
        context['condition8'] = self.condition8
        context['condition9'] = self.condition9
        context['condition10'] = self.condition10
        context['condition11'] = self.condition11
        context['condition12'] = self.condition12
        context['condition13'] = self.condition13
        context['condition14'] = self.condition14
        context['condition15'] = self.condition15
        context['condition16'] = self.condition16
        context['condition17'] = self.condition17
        context['condition18'] = self.condition18
        context['condition19'] = self.condition19
        context['condition20'] = self.condition20

        context['loss_rate'] = ""

        # # todo
        loss_rate= DelConInfo.objects.values('loss_rate').first()
        loss_status = DelConInfo.objects.values('loss_rate_status').first()
        if loss_status['loss_rate_status'] == '计算完成':
            DelConInfo.objects.filter(loss_rate_status="计算完成").update(loss_rate_status="未计算")
            context['loss_rate'] = loss_rate['loss_rate']

        return context

def upload_bank(request):
    if request.method == "POST":  # 验证POST
        # todo 获取文件名称
        excel_name = request.FILES['my_file']
        wb = load_workbook(filename=excel_name)
        ws = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws[0])
        max_row = ws.max_row
        max_col = ws.max_column
        excel_name = str(excel_name)

        # todo 保存表信息
        table_info = GeneralTableInfo()
        table_info.table_name = excel_name
        table_info.col_number = max_col
        table_info.col_1 = ws.cell(row=1, column=1).value
        table_info.col_2 = ws.cell(row=1, column=2).value
        table_info.col_3 = ws.cell(row=1, column=3).value
        table_info.col_4 = ws.cell(row=1, column=4).value
        table_info.col_5 = ws.cell(row=1, column=5).value
        table_info.col_6 = ws.cell(row=1, column=6).value
        table_info.col_7 = ws.cell(row=1, column=7).value
        table_info.col_8 = ws.cell(row=1, column=8).value
        table_info.col_9 = ws.cell(row=1, column=9).value
        table_info.col_10 = ws.cell(row=1, column=10).value
        table_info.col_11 = ws.cell(row=1, column=11).value
        table_info.col_12 = ws.cell(row=1, column=12).value
        table_info.col_13 = ws.cell(row=1, column=13).value
        table_info.col_14 = ws.cell(row=1, column=14).value
        table_info.col_15 = ws.cell(row=1, column=15).value
        table_info.col_16 = ws.cell(row=1, column=16).value
        table_info.col_17 = ws.cell(row=1, column=17).value
        table_info.col_18 = ws.cell(row=1, column=18).value
        table_info.col_19 = ws.cell(row=1, column=19).value
        table_info.col_20 = ws.cell(row=1, column=20).value

        is_there = GeneralTableInfo.objects.filter(table_name=excel_name)
        if is_there:
            None
        else:
            table_info.save()

        # 删除信息
        is_delete_save = GeneralMaskInfo.objects.filter(table_name=excel_name)
        if is_delete_save:
            None
        else:
            mask = GeneralMaskInfo()
            mask.table_name = excel_name
            mask.col_1 = "1"
            mask.col_2 = "1"
            mask.col_3 = "1"
            mask.col_4 = "1"
            mask.col_5 = "1"
            mask.col_6 = "1"
            mask.col_7 = "1"
            mask.col_8 = "1"
            mask.col_9 = "1"
            mask.col_10 = "1"
            mask.col_11 = "1"
            mask.col_12 = "1"
            mask.col_13 = "1"
            mask.col_14 = "1"
            mask.col_15 = "1"
            mask.col_16 = "1"
            mask.col_17 = "1"
            mask.col_18 = "1"
            mask.col_19 = "1"
            mask.col_20 = "1"
            mask.save()



        # todo 保存数据信息
        for row in range(2, max_row + 1):
            # 获取表单元素
            data_info = GeneralInfo()
            data_info.table_name = excel_name
            data_info.col_1 = ws.cell(row=row, column=1).value
            data_info.col_2 = ws.cell(row=row, column=2).value
            data_info.col_3 = ws.cell(row=row, column=3).value
            data_info.col_4 = ws.cell(row=row, column=4).value
            data_info.col_5 = ws.cell(row=row, column=5).value
            data_info.col_6 = ws.cell(row=row, column=6).value
            data_info.col_7 = ws.cell(row=row, column=7).value
            data_info.col_8 = ws.cell(row=row, column=8).value
            data_info.col_9 = ws.cell(row=row, column=9).value
            data_info.col_10 = ws.cell(row=row, column=10).value
            data_info.col_11 = ws.cell(row=row, column=11).value
            data_info.col_12 = ws.cell(row=row, column=12).value
            data_info.col_13 = ws.cell(row=row, column=13).value
            data_info.col_14 = ws.cell(row=row, column=14).value
            data_info.col_15 = ws.cell(row=row, column=15).value
            data_info.col_16 = ws.cell(row=row, column=16).value
            data_info.col_17 = ws.cell(row=row, column=17).value
            data_info.col_18 = ws.cell(row=row, column=18).value
            data_info.col_19 = ws.cell(row=row, column=19).value
            data_info.col_20 = ws.cell(row=row, column=20).value

            data_info.save()
        # todo 当下正在使用的表
        ParameterInfo.objects.filter(id=1).update(fill_name=excel_name)
        messages.success(request, "数据导入成功")
        return redirect('upload_bank')
    return redirect('upload_bank')

# todo 数据归一化处理
def bankinfo_nov(request):
    all_banks = BankInfo.objects.values('cash_central_bank_funds',
                                        'sellable_assetset',
                                        'risk_preparation',
                                        'interest_payable',
                                        'paid_in_capital',
                                        'fixed_assets',
                                        'total_assets',
                                        'capital_reserves')
    all_banks_id =  BankInfo.objects.values('id')
    df = pd.DataFrame(list(all_banks))
    df_id = pd.DataFrame(list(all_banks_id))
    df_arr_id = np.array(df_id)

    # todo 文本类型转换
    df['cash_central_bank_funds'] = pd.to_numeric(df['cash_central_bank_funds'])
    df['sellable_assetset'] = pd.to_numeric(df['sellable_assetset'])
    df['risk_preparation'] = pd.to_numeric(df['risk_preparation'])
    df['interest_payable'] = pd.to_numeric(df['interest_payable'])
    df['paid_in_capital'] = pd.to_numeric(df['paid_in_capital'])
    df['fixed_assets'] = pd.to_numeric(df['fixed_assets'])
    df['total_assets'] = pd.to_numeric(df['total_assets'])
    df['capital_reserves'] = pd.to_numeric(df['capital_reserves'])

    df_arr = np.array(df)
    df_nor = normalization(df_arr)
    i = 0
    for nor in df_nor:
        BankInfo.objects.filter(id=df_arr_id[i]).update(cash_central_bank_funds=round(nor[0],6))
        BankInfo.objects.filter(id=df_arr_id[i]).update(sellable_assetset =round(nor[1],6))
        BankInfo.objects.filter(id=df_arr_id[i]).update(risk_preparation = round(nor[2],6))
        BankInfo.objects.filter(id=df_arr_id[i]).update(interest_payable = round(nor[3],6))
        BankInfo.objects.filter(id=df_arr_id[i]).update(paid_in_capital = round(nor[4],6))
        BankInfo.objects.filter(id=df_arr_id[i]).update(fixed_assets = round(nor[5],6))
        BankInfo.objects.filter(id=df_arr_id[i]).update(total_assets = round(nor[6],6))
        BankInfo.objects.filter(id=df_arr_id[i]).update(capital_reserves = round(nor[7],6))
        i = i + 1

    messages.success(request, "数据归一化处理成功")
    return redirect('data_nov')


# todo 计算缺失率
def bankinfo_con(request):

    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_name = fill_name[0]['fill_name']

    coll_1 = GeneralInfo.objects.filter(
        Q(col_1__isnull=True) | Q(col_1__exact=''), table_name=fill_name)
    coll_2 = GeneralInfo.objects.filter(
        Q(col_2__isnull=True) | Q(col_2__exact=''), table_name=fill_name)
    coll_3 = GeneralInfo.objects.filter(
        Q(col_3__isnull=True) | Q(col_3__exact=''), table_name=fill_name)
    coll_4 = GeneralInfo.objects.filter(
        Q(col_4__isnull=True) | Q(col_4__exact=''), table_name=fill_name)
    coll_5 = GeneralInfo.objects.filter(
        Q(col_5__isnull=True) | Q(col_5__exact=''), table_name=fill_name)
    coll_6 = GeneralInfo.objects.filter(
        Q(col_6__isnull=True) | Q(col_6__exact=''), table_name=fill_name)
    coll_7 = GeneralInfo.objects.filter(
        Q(col_7__isnull=True) | Q(col_7__exact=''), table_name=fill_name)
    coll_8 = GeneralInfo.objects.filter(
        Q(col_8__isnull=True) | Q(col_8__exact=''), table_name=fill_name)
    coll_9 = GeneralInfo.objects.filter(
        Q(col_9__isnull=True) | Q(col_9__exact=''), table_name=fill_name)
    coll_10 = GeneralInfo.objects.filter(
        Q(col_10__isnull=True) | Q(col_10__exact=''), table_name=fill_name)
    coll_11 = GeneralInfo.objects.filter(
        Q(col_11__isnull=True) | Q(col_11__exact=''), table_name=fill_name)
    coll_12 = GeneralInfo.objects.filter(
        Q(col_12__isnull=True) | Q(col_12__exact=''), table_name=fill_name)
    coll_13 = GeneralInfo.objects.filter(
        Q(col_13__isnull=True) | Q(col_13__exact=''), table_name=fill_name)
    coll_14 = GeneralInfo.objects.filter(
        Q(col_14__isnull=True) | Q(col_14__exact=''), table_name=fill_name)
    coll_15 = GeneralInfo.objects.filter(
        Q(col_15__isnull=True) | Q(col_15__exact=''), table_name=fill_name)
    coll_16 = GeneralInfo.objects.filter(
        Q(col_16__isnull=True) | Q(col_16__exact=''), table_name=fill_name)
    coll_17 = GeneralInfo.objects.filter(
        Q(col_17__isnull=True) | Q(col_17__exact=''), table_name=fill_name)
    coll_18 = GeneralInfo.objects.filter(
        Q(col_18__isnull=True) | Q(col_18__exact=''), table_name=fill_name)
    coll_19 = GeneralInfo.objects.filter(
        Q(col_19__isnull=True) | Q(col_19__exact=''), table_name=fill_name)
    coll_20 = GeneralInfo.objects.filter(
        Q(col_20__isnull=True) | Q(col_20__exact=''), table_name=fill_name)

    all_banks = GeneralInfo.objects.filter(table_name=fill_name).all()
    col_num = GeneralTableInfo.objects.filter(table_name=fill_name).values('col_number').first()
    col_num = col_num['col_number']
    all_count = all_banks.count() * col_num
    loss_count_list = [coll_1.count(), coll_2.count(), coll_3.count(),coll_4.count(),
                  coll_5.count(), coll_6.count(), coll_7.count(),coll_8.count(),
                  coll_9.count(), coll_10.count(), coll_11.count(),coll_12.count(),
                  coll_13.count(), coll_14.count(), coll_15.count(),coll_16.count(),
                  coll_17.count(), coll_18.count(), coll_19.count(),coll_20.count()]

    loss_count = 0
    for i in range(col_num):
        loss_count = loss_count +loss_count_list[i]

    loss_rate = round(loss_count / all_count, 2)
    loss_id = DelConInfo.objects.values('id').first()
    DelConInfo.objects.filter(id=loss_id['id']).update(loss_rate=loss_rate)
    DelConInfo.objects.filter(id=loss_id['id']).update(loss_rate_status="计算完成")

    messages.success(request, "计算缺失率成功")
    return redirect('data_con')




# todo 构建模型
class ModelBuildView(LoginRequiredMixin,ListView):
    login_url = 'login'

    model = GeneralInfo
    context_object_name = 'books'
    template_name = 'book/model_build_general.html'
    search_value = ""
    order_field = "id"

    # todo
    number_rmse = []
    number = []
    # todo 新加控制属性
    # search_int = 0
    # DelConInfo.objects.filter(del_status__icontains="完成").delete()
    condition1 = ""
    condition2 = ""
    condition3 = ""
    condition4 = ""
    condition5 = ""
    condition6 = ""
    condition7 = ""
    condition8 = ""
    condition9 = ""
    condition10 = ""
    condition11 = ""
    condition12 = ""
    condition13 = ""
    condition14 = ""
    condition15 = ""
    condition16 = ""
    condition17 = ""
    condition18 = ""
    condition19 = ""
    condition20 = ""

    def get_queryset(self):
        search = self.request.GET.get("search")
        search_q = self.request.GET.get("search_q")

        # todo 加一个判断，：当下正在操作的表
        fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
        fill_nameed = fill_name[0]['fill_name']

        self.condition1 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        self.condition2 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        self.condition3 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        self.condition4 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        self.condition5 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        self.condition6 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        self.condition7 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        self.condition8 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        self.condition9 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        self.condition10 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        self.condition11 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        self.condition12 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        self.condition13 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        self.condition14 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        self.condition15 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        self.condition16 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        self.condition17 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        self.condition18 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        self.condition19 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        self.condition20 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        status_con1 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        status_con2 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        status_con3 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        status_con4 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        status_con5 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        status_con6 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        status_con7 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        status_con8 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        status_con9 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        status_con10 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        status_con11 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        status_con12 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        status_con13 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        status_con14 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        status_con15 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        status_con16 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        status_con17 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        status_con18 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        status_con19 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        status_con20 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        if status_con1 == "2":
            self.condition1 = ""
        if status_con2 == "2":
            self.condition2 = ""
        if status_con3 == "2":
            self.condition3 = ""
        if status_con4 == "2":
            self.condition4 = ""
        if status_con5 == "2":
            self.condition5 = ""
        if status_con6 == "2":
            self.condition6 = ""
        if status_con7 == "2":
            self.condition7 = ""
        if status_con8 == "2":
            self.condition8 = ""
        if status_con9 == "2":
            self.condition9 = ""
        if status_con10 == "2":
            self.condition10 = ""
        if status_con11 == "2":
            self.condition11 = ""
        if status_con12 == "2":
            self.condition12 = ""
        if status_con13 == "2":
            self.condition13 = ""
        if status_con14 == "2":
            self.condition14 = ""
        if status_con15 == "2":
            self.condition15 = ""
        if status_con16 == "2":
            self.condition16 = ""
        if status_con17 == "2":
            self.condition17 = ""
        if status_con18 == "2":
            self.condition18 = ""
        if status_con19 == "2":
            self.condition19 = ""
        if status_con20 == "2":
            self.condition20 = ""
        all_books = GeneralInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)

        if search and search_q:

            all_banks = GeneralInfo.objects.filter(table_name=fill_nameed).values('col_1',
                                                                                  'col_2',
                                                                                  'col_3',
                                                                                  'col_4',
                                                                                  'col_5',
                                                                                  'col_6',
                                                                                  'col_7',
                                                                                  'col_8',
                                                                                  'col_9',
                                                                                  'col_10',
                                                                                  'col_11',
                                                                                  'col_12',
                                                                                  'col_13',
                                                                                  'col_14',
                                                                                  'col_15',
                                                                                  'col_16',
                                                                                  'col_17',
                                                                                  'col_18',
                                                                                  'col_19',
                                                                                  'col_20',).order_by(self.order_field)

            search_int = int(search)
            search_q_f = float(search_q)
            # todo 实现数据转化填补，先指定
            df = pd.DataFrame(list(all_banks))

            # 取
            col_num = GeneralTableInfo.objects.filter(table_name=fill_nameed).values('col_number').first()
            col_num = col_num['col_number']

            df = df.iloc[:, 0:col_num]

            for i in range(col_num):
                df['col_'+str(i+1)] = pd.to_numeric(df['col_'+str(i+1)])

            df_narry = np.array(df)
            print(df_narry)
            _, _, rmse_list, theta_G = run_model(df_narry, search_int, search_q_f)
            list_x = [i for i in range(1, search_int + 1)]

            # todo 保存模型参数
            modelpa = ModelParameter.objects.all()
            modelpa.filter(id=1).update(model_G=theta_G)

            self.number = list_x
            self.number_rmse = rmse_list

        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)

        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(ModelBuildView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()

        # todo 放在上一个方法里会执行 两次
        # todo
        context['condition1'] = self.condition1
        context['condition2'] = self.condition2
        context['condition3'] = self.condition3
        context['condition4'] = self.condition4
        context['condition5'] = self.condition5
        context['condition6'] = self.condition6
        context['condition7'] = self.condition7
        context['condition8'] = self.condition8
        context['condition9'] = self.condition9
        context['condition10'] = self.condition10
        context['condition11'] = self.condition11
        context['condition12'] = self.condition12
        context['condition13'] = self.condition13
        context['condition14'] = self.condition14
        context['condition15'] = self.condition15
        context['condition16'] = self.condition16
        context['condition17'] = self.condition17
        context['condition18'] = self.condition18
        context['condition19'] = self.condition19
        context['condition20'] = self.condition20

        context['number'] = self.number
        context['number_rmse'] = self.number_rmse


        return context


# todo 清空数据
def data_clear(request):
    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_name = fill_name[0]['fill_name']
    all_banks = GeneralInfo.objects.filter(table_name=fill_name).all()

    all_banks.delete()
    messages.success(request, "数据已清空")
    return redirect('model_building')

# todo 改为导出数据
def output_save(request):
    # f = xlwt.Workbook(encoding = 'utf-8')

    # todo 当前操作的表
    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_nameed = fill_name[0]['fill_name']

    excel_name = fill_nameed.split('.')[0]+str(datetime.now()).split('.')[-1]

    f = xlsxwriter.Workbook(f'C:/Users/80669/Desktop/填补系统数据/%s.xlsx'%(excel_name))
    sheet1 = f.add_worksheet('导出数据')

    condition1 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
    condition2 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
    condition3 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
    condition4 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
    condition5 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
    condition6 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
    condition7 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
    condition8 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
    condition9 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
    condition10 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
    condition11 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
    condition12 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
    condition13 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
    condition14 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
    condition15 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
    condition16 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
    condition17 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
    condition18 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
    condition19 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
    condition20 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']


    col_num = GeneralTableInfo.objects.filter(table_name=fill_nameed).values('col_number').first()
    col_num = col_num['col_number']
    col_list = [condition1, condition2, condition3, condition4,
                condition5, condition6, condition7, condition8,
                condition9, condition10, condition11, condition12,
                condition13, condition14, condition15, condition16,
                condition17, condition18, condition19, condition20,]

    # todo 需要删除表判断下
    status_con1 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
    status_con2 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
    status_con3 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
    status_con4 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
    status_con5 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
    status_con6 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
    status_con7 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
    status_con8 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
    status_con9 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
    status_con10 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
    status_con11 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
    status_con12 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
    status_con13 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
    status_con14 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
    status_con15 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
    status_con16 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
    status_con17 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
    status_con18 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
    status_con19 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
    status_con20 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

    del_list = [status_con1,status_con2,status_con3,status_con4,status_con5,
                status_con6,status_con7,status_con8,status_con9,status_con10,
                status_con11,status_con12,status_con13,status_con14,status_con15,
                status_con16,status_con17,status_con18,status_con19,status_con20]

    # Excel表的第一行
    row0 = []
    for i in range(col_num):
        if del_list[i] != "2":
            row0.append(col_list[i])

    for m in range(0, len(row0)):
        sheet1.write(0, m, row0[m])

    temp = []
    queryset = GeneralInfo.objects.filter(table_name=fill_nameed).all()
    for d in queryset:
        t = [d.col_1, d.col_2, d.col_3, d.col_4, d.col_5,
             d.col_6, d.col_7, d.col_8, d.col_9, d.col_10,
             d.col_11, d.col_12, d.col_13, d.col_14, d.col_15,
             d.col_16, d.col_17, d.col_18, d.col_19, d.col_20]
        t_temp = []
        for i in range(col_num):
            if del_list[i] != "2":
                t_temp.append(t[i])
        # 将数据保存在temp数组中
        temp.append(t_temp)
    i = 1
    for t in temp:
        for j in range(len(t)):
            # 数据按行列一次导入到Excel表中
            sheet1.write(i, j, t[j])
        i = i + 1

    f.close()

    messages.success(request, "数据导出成功")
    # modelpa = ModelParameter.objects.values('model_G').first()
    # todo str类型
    return redirect('model_building')



# todo 导出填补数据
def filled_save(request):

    # todo 当前操作的表
    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_nameed = fill_name[0]['fill_name']

    excel_name = fill_nameed.split('.')[0]+str(datetime.now()).split('.')[-1] +"填补后数据"

    f = xlsxwriter.Workbook(f'C:/Users/80669/Desktop/填补系统数据/%s.xlsx'%(excel_name))
    sheet1 = f.add_worksheet(fill_nameed+'填补后数据')

    condition1 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
    condition2 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
    condition3 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
    condition4 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
    condition5 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
    condition6 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
    condition7 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
    condition8 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
    condition9 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
    condition10 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
    condition11 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
    condition12 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
    condition13 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
    condition14 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
    condition15 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
    condition16 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
    condition17 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
    condition18 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
    condition19 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
    condition20 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']


    col_num = GeneralTableInfo.objects.filter(table_name=fill_nameed).values('col_number').first()
    col_num = col_num['col_number']
    col_list = [condition1, condition2, condition3, condition4,
                condition5, condition6, condition7, condition8,
                condition9, condition10, condition11, condition12,
                condition13, condition14, condition15, condition16,
                condition17, condition18, condition19, condition20,]

    # Excel表的第一行
    row0 = []
    for i in range(col_num):
        row0.append(col_list[i])

    for m in range(0, len(row0)):
        sheet1.write(0, m, row0[m])

    temp = []
    queryset = GeneralFilledInfo.objects.filter(table_name=fill_nameed).all()
    for d in queryset:
        t = [d.col_1, d.col_2, d.col_3, d.col_4, d.col_5,
             d.col_6, d.col_7, d.col_8, d.col_9, d.col_10,
             d.col_11, d.col_12, d.col_13, d.col_14, d.col_15,
             d.col_16, d.col_17, d.col_18, d.col_19, d.col_20]
        t_temp = []
        for i in range(col_num):
            t_temp.append(t[i])
        # 将数据保存在temp数组中
        temp.append(t_temp)
    i = 1
    for t in temp:
        for j in range(len(t)):
            # 数据按行列一次导入到Excel表中
            sheet1.write(i, j, t[j])
        i = i + 1

    f.close()

    messages.success(request, "填补数据导出成功")
    # modelpa = ModelParameter.objects.values('model_G').first()
    # todo str类型
    return redirect('fill_data')

# todo 填补数据before
class FillDataView(LoginRequiredMixin,ListView):
    login_url = 'login'
    model = BankInfo
    context_object_name = 'books'
    template_name = 'book/fill_data.html'
    search_value = ""
    order_field = "id"

    # todo
    number_rmse = []
    number = []
    # todo 新加控制属性
    message_do = 1
    # search_int = 0
    # DelConInfo.objects.filter(del_status__icontains="完成").delete()
    condition1 = True
    condition2 = True
    condition3 = True
    condition4 = True
    condition5 = True
    condition6 = True
    condition7 = True
    condition8 = True
    condition9 = True
    condition10 = True
    condition11 = True
    condition12 = True

    def get_queryset(self):
        search = self.request.GET.get("search")

        status_con1 = BankInfoDelMark.objects.values("company_name").first()
        status_con2 = BankInfoDelMark.objects.values("type").first()
        status_con3 = BankInfoDelMark.objects.values("bbd_url").first()
        status_con4 = BankInfoDelMark.objects.values("bbd_type").first()
        status_con5 = BankInfoDelMark.objects.values("cash_central_bank_funds").first()
        status_con6 = BankInfoDelMark.objects.values("sellable_assetset").first()
        status_con7 = BankInfoDelMark.objects.values("risk_preparation").first()
        status_con8 = BankInfoDelMark.objects.values("interest_payable").first()
        status_con9 = BankInfoDelMark.objects.values("paid_in_capital").first()
        status_con10 = BankInfoDelMark.objects.values("fixed_assets").first()
        status_con11 = BankInfoDelMark.objects.values("total_assets").first()
        status_con12 = BankInfoDelMark.objects.values("capital_reserves").first()

        if status_con1['company_name'] == "2":
            self.condition1 = False
        if status_con2['type'] == "2":
            self.condition2 = False
        if status_con3['bbd_url'] == "2":
            self.condition3 = False
        if status_con4['bbd_type'] == "2":
            self.condition4 = False
        if status_con5['cash_central_bank_funds'] == "2":
            self.condition5 = False
        if status_con6['sellable_assetset'] == "2":
            self.condition6 = False
        if status_con7['risk_preparation'] == "2":
            self.condition7 = False
        if status_con8['interest_payable'] == "2":
            self.condition8 = False
        if status_con9['paid_in_capital'] == "2":
            self.condition9 = False
        if status_con10['fixed_assets'] == "2":
            self.condition10 = False
        if status_con11['total_assets'] == "2":
            self.condition11 = False
        if status_con12['capital_reserves'] == "2":
            self.condition12 = False

        # if order_by:
        #     all_books = BankInfo.objects.all().order_by(order_by)
        #     self.order_field = order_by
        # else:
        all_books = BankInfo.objects.all().order_by(self.order_field)
        # todo 这是搜索框条件，如果有内容执行下列程序
        if search:

            all_banks = BankInfo.objects.values('cash_central_bank_funds',
                                                'sellable_assetset',
                                                'risk_preparation',
                                                'interest_payable',
                                                'paid_in_capital',
                                                'fixed_assets',
                                                'total_assets',
                                                'capital_reserves')
            search_int = int(search)
            # todo 实现数据转化填补，先指定
            df = pd.DataFrame(list(all_banks))
            # todo 文本类型转数字
            df['cash_central_bank_funds'] = pd.to_numeric(df['cash_central_bank_funds'])
            df['sellable_assetset'] = pd.to_numeric(df['sellable_assetset'])
            df['risk_preparation'] = pd.to_numeric(df['risk_preparation'])
            df['interest_payable'] = pd.to_numeric(df['interest_payable'])
            df['paid_in_capital'] = pd.to_numeric(df['paid_in_capital'])
            df['fixed_assets'] = pd.to_numeric(df['fixed_assets'])
            df['total_assets'] = pd.to_numeric(df['total_assets'])
            df['capital_reserves'] = pd.to_numeric(df['capital_reserves'])
            df_narry = np.array(df)
            _, _, rmse_list, theta_G = run_model(df_narry, search_int)
            list_x = [i for i in range(1, search_int+1)]

            self.number = list_x
            self.number_rmse = rmse_list

            # todo 保存模型参数


            # all_books = all_books.filter(
            #     Q(id__icontains=search) | Q(username__icontains=search)
            # )
            # status_del = DelConInfo.objects.values("del_name").last()
            # print(status_del['del_name'])
        else:
            self.message_do = 1

        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)
        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(FillDataView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['search'] = self.search_value
        context['orderby'] = self.order_field
        context['objects'] = self.get_queryset()

        # todo 放在上一个方法里会执行 两次
        if self.message_do == 2:
            messages.warning(self.request, f"没有")
        # todo
        context['condition1'] = self.condition1
        context['condition2'] = self.condition2
        context['condition3'] = self.condition3
        context['condition4'] = self.condition4
        context['condition5'] = self.condition5
        context['condition6'] = self.condition6
        context['condition7'] = self.condition7
        context['condition8'] = self.condition8
        context['condition9'] = self.condition9
        context['condition10'] = self.condition10
        context['condition11'] = self.condition11
        context['condition12'] = self.condition12
        context['number'] = self.number
        context['number_rmse'] = self.number_rmse

        return context


# todo 填补数据
class FillDataView(LoginRequiredMixin,ListView):
    login_url = 'login'

    model = GeneralInfo
    context_object_name = 'books'
    template_name = 'book/fill_data_general.html'
    search_value = ""
    order_field = "id"

    # todo
    number_rmse = []
    number = []
    # todo 新加控制属性
    message_do = 1
    condition1 = ""
    condition2 = ""
    condition3 = ""
    condition4 = ""
    condition5 = ""
    condition6 = ""
    condition7 = ""
    condition8 = ""
    condition9 = ""
    condition10 = ""
    condition11 = ""
    condition12 = ""
    condition13 = ""
    condition14 = ""
    condition15 = ""
    condition16 = ""
    condition17 = ""
    condition18 = ""
    condition19 = ""
    condition20 = ""

    def get_queryset(self):
        # todo 加一个判断：当下正在操作的表
        fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
        fill_nameed = fill_name[0]['fill_name']

        self.condition1 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        self.condition2 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        self.condition3 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        self.condition4 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        self.condition5 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        self.condition6 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        self.condition7 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        self.condition8 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        self.condition9 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        self.condition10 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        self.condition11 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        self.condition12 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        self.condition13 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        self.condition14 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        self.condition15 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        self.condition16 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        self.condition17 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        self.condition18 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        self.condition19 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        self.condition20 = GeneralTableInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        # search = self.request.GET.get("search")
        status_con1 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_1").first()['col_1']
        status_con2 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_2").first()['col_2']
        status_con3 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_3").first()['col_3']
        status_con4 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_4").first()['col_4']
        status_con5 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_5").first()['col_5']
        status_con6 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_6").first()['col_6']
        status_con7 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_7").first()['col_7']
        status_con8 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_8").first()['col_8']
        status_con9 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_9").first()['col_9']
        status_con10 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_10").first()['col_10']
        status_con11 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_11").first()['col_11']
        status_con12 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_12").first()['col_12']
        status_con13 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_13").first()['col_13']
        status_con14 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_14").first()['col_14']
        status_con15 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_15").first()['col_15']
        status_con16 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_16").first()['col_16']
        status_con17 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_17").first()['col_17']
        status_con18 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_18").first()['col_18']
        status_con19 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_19").first()['col_19']
        status_con20 = GeneralMaskInfo.objects.filter(table_name=fill_nameed).values("col_20").first()['col_20']

        if status_con1 == "2":
            self.condition1 = ""
        if status_con2 == "2":
            self.condition2 = ""
        if status_con3 == "2":
            self.condition3 = ""
        if status_con4 == "2":
            self.condition4 = ""
        if status_con5 == "2":
            self.condition5 = ""
        if status_con6 == "2":
            self.condition6 = ""
        if status_con7 == "2":
            self.condition7 = ""
        if status_con8 == "2":
            self.condition8 = ""
        if status_con9 == "2":
            self.condition9 = ""
        if status_con10 == "2":
            self.condition10 = ""
        if status_con11 == "2":
            self.condition11 = ""
        if status_con12 == "2":
            self.condition12 = ""
        if status_con13 == "2":
            self.condition13 = ""
        if status_con14 == "2":
            self.condition14 = ""
        if status_con15 == "2":
            self.condition15 = ""
        if status_con16 == "2":
            self.condition16 = ""
        if status_con17 == "2":
            self.condition17 = ""
        if status_con18 == "2":
            self.condition18 = ""
        if status_con19 == "2":
            self.condition19 = ""
        if status_con20 == "2":
            self.condition20 = ""
        all_books = GeneralInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)

        is_filled = FilledParameter.objects.values('status').last()
        self.rmse = FilledParameter.objects.values('rmse').last()['rmse']
        self.all_books_filled = []
        self.count_total_filled = 0
        if is_filled[('status')] == "2":
            self.all_books_filled = GeneralFilledInfo.objects.filter(table_name=fill_nameed).all().order_by(self.order_field)
            self.count_total_filled = self.all_books_filled.count()
            paginator = Paginator(self.all_books_filled, PAGINATOR_NUMBER)
            page = self.request.GET.get('page')
            self.all_books_filled = paginator.get_page(page)

        self.count_total = all_books.count()
        paginator = Paginator(all_books, PAGINATOR_NUMBER)
        page = self.request.GET.get('page')
        books = paginator.get_page(page)

        return books

    def get_context_data(self, *args, **kwargs):
        # TODO 这个要注意hmf
        context = super(FillDataView, self).get_context_data(*args, **kwargs)
        # context = {}
        context['count_total'] = self.count_total
        context['count_total_filled'] = self.count_total_filled
        context['objects'] = self.get_queryset()

        context['search'] = ""
        context['orderby'] = ""

        # todo 放在上一个方法里会执行 两次
        # todo
        context['condition1'] = self.condition1
        context['condition2'] = self.condition2
        context['condition3'] = self.condition3
        context['condition4'] = self.condition4
        context['condition5'] = self.condition5
        context['condition6'] = self.condition6
        context['condition7'] = self.condition7
        context['condition8'] = self.condition8
        context['condition9'] = self.condition9
        context['condition10'] = self.condition10
        context['condition11'] = self.condition11
        context['condition12'] = self.condition12
        context['condition13'] = self.condition13
        context['condition14'] = self.condition14
        context['condition15'] = self.condition15
        context['condition16'] = self.condition16
        context['condition17'] = self.condition17
        context['condition18'] = self.condition18
        context['condition19'] = self.condition19
        context['condition20'] = self.condition20

        context['books_fill'] = self.all_books_filled
        context['rmse'] = self.rmse

        return context


# todo 填补数据
def data_fill(request):

    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_name = fill_name[0]['fill_name']

    all_banks = GeneralInfo.objects.filter(table_name=fill_name).values('col_1',
                                                                        'col_2',
                                                                        'col_3',
                                                                        'col_4',
                                                                        'col_5',
                                                                        'col_6',
                                                                        'col_7',
                                                                        'col_8',
                                                                        'col_9',
                                                                        'col_10',
                                                                        'col_11',
                                                                        'col_12',
                                                                        'col_13',
                                                                        'col_14',
                                                                        'col_15',
                                                                        'col_16',
                                                                        'col_17',
                                                                        'col_18',
                                                                        'col_19',
                                                                        'col_20',).order_by("id")

    search_int = 100
    # todo 实现数据转化填补，先指定
    df = pd.DataFrame(list(all_banks))

    # 取
    col_num = GeneralTableInfo.objects.filter(table_name=fill_name).values('col_number').first()
    col_num = col_num['col_number']

    df = df.iloc[:, 0:col_num]

    for i in range(col_num):
        df['col_' + str(i + 1)] = pd.to_numeric(df['col_' + str(i + 1)])
    df_narry = np.array(df)

    data_filled, rmse = datafill(df_narry, search_int)
    for i in range(len(df_narry)):
        upload_bank = GeneralFilledInfo()
        upload_bank.table_name = fill_name
        upload_bank.col_1 = round(data_filled[i][0], 4)
        upload_bank.col_2 = round(data_filled[i][1], 4)
        upload_bank.col_3 = round(data_filled[i][2], 4)
        if 3 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_4 = round(data_filled[i][3], 4)
        if 4 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_5 = round(data_filled[i][4], 4)
        if 5 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_6 = round(data_filled[i][5], 4)
        if 6 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_7 = round(data_filled[i][6], 4)
        if 7 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_8 = round(data_filled[i][7], 4)
        if 8 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_9 = round(data_filled[i][8], 4)
        if 9 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_10 = round(data_filled[i][9], 4)
        if 10 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_11 = round(data_filled[i][10], 4)
        if 11 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_12 = round(data_filled[i][11], 4)
        if 12 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_13 = round(data_filled[i][12], 4)
        if 13 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_14 = round(data_filled[i][13], 4)
        if 14 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_15 = round(data_filled[i][14], 4)
        if 15 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_16 = round(data_filled[i][15], 4)
        if 16 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_17 = round(data_filled[i][16], 4)
        if 17 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_18 = round(data_filled[i][17], 4)
        if 18 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_19 = round(data_filled[i][18], 4)
        if 19 == col_num:
            upload_bank.save()
            continue
        upload_bank.col_20 = round(data_filled[i][19], 4)

        upload_bank.save()
    FilledParameter.objects.filter(id=1).update(status="2")
    FilledParameter.objects.filter(id=1).update(rmse=rmse)

    messages.success(request, "数据填补已经完成")
    return redirect('fill_data')

# todo 清空数据-填补前专用
def data_clear_fill(request):
    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_name = fill_name[0]['fill_name']
    all_banks = GeneralInfo.objects.filter(table_name=fill_name).all()

    all_banks.delete()
    messages.success(request, "数据已清空")
    return redirect('fill_data')

# todo 清空数据-填补后专用
def dataclear(request):
    fill_name = ParameterInfo.objects.filter(id=1).values('fill_name')
    fill_name = fill_name[0]['fill_name']

    all_banks = GeneralFilledInfo.objects.filter(table_name=fill_name).all()

    all_banks.delete()
    FilledParameter.objects.filter(id=1).update(status="1")
    FilledParameter.objects.filter(id=1).update(rmse="")
    messages.success(request, "填补数据已清空")
    return redirect('fill_data')


# todo 保存填补后数据
def fill_save(request):
    messages.success(request, "数据已保存")
    # modelpa = ModelParameter.objects.values('model_G').first()
    # todo str类型
    return redirect('fill_data')


#     todo 权限展示

def power_show(request):
    messages.err(request, "此用户没有该功能权限")
    return render(request, "book/show.html")
